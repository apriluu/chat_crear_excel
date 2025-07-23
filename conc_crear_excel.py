import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog
import pandas as pd
import json
from openpyxl import load_workbook

def afegir_json_a_excel():
    text = text_area.get("1.0", tk.END).strip()
    if not text:
        messagebox.showwarning("Atenció", "El camp de text està buit.")
        return

    try:
        data = json.loads(text)

        def llista_a_text(llista):
            if isinstance(llista, list):
                return "\n".join(map(str, llista))
            elif isinstance(llista, str):
                return llista.strip()
            elif llista is None:
                return ""
            else:
                return str(llista)

        def perfils_tecnics_a_text(perfils):
            if isinstance(perfils, dict):
                perfils = [perfils]
            if not perfils or not isinstance(perfils, list):
                return ""

            resultat = []
            for perfil in perfils:
                if not isinstance(perfil, dict):
                    continue
                titol = perfil.get("titulacio", "")
                experiencia = perfil.get("experiencia", "")
                funcions = perfil.get("funcions", "")
                resultat.append(f"Títol: {titol}\nExperiència: {experiencia}\nFuncions: {funcions}")
            return "\n\n".join(resultat)

        equivalencies = {
            "nom_projecte": ["nom_projecte", "nom_del_projecte"],
            "ubicacio_projecte": ["ubicacio_projecte", "ubicacio_del_projecte"],
            "pressupost_licitacio": ["pressupost_licitacio", "PEM", "pressupost_licitacio_PEM",
                                     "pressupost_de_licitacio_PEM"],
            "data_licitacio": ["data_licitacio", "data_de_licitacio"],
            "perfils_tecnics_requerits": ["perfils_tecnics_requerits", "equips_tecnics"],
            "termini_execucio": ["termini_execucio", "termini_d_execucio"],
            "requisits_legals_tecnics": ["requisits_legals_tecnics_destacats", "requisits_legals_o_tecnics_destacats"],
            "documentacio": ["documentacio_a_aportar", "documentacio_a_portar", "documentacio_aportar",
                             "documents_a_presentar"]
        }

        def get_valor_equivalent(data, claus):
            for clau in claus:
                if clau in data:
                    return data[clau]
            return None

        dades_planes = {
            "Nom del projecte": get_valor_equivalent(data, equivalencies["nom_projecte"]),
            "Ubicació": get_valor_equivalent(data, equivalencies["ubicacio_projecte"]),
            "Pressupost": get_valor_equivalent(data, equivalencies["pressupost_licitacio"]),
            "Data licitació": get_valor_equivalent(data, equivalencies["data_licitacio"]),
            "Perfils tècnics": perfils_tecnics_a_text(
                get_valor_equivalent(data, equivalencies["perfils_tecnics_requerits"])),
            "Termini execució": get_valor_equivalent(data, equivalencies["termini_execucio"]),

            "Requisits legals/tècnics": llista_a_text(
                get_valor_equivalent(data, equivalencies["requisits_legals_tecnics"])),
            "Documentacio": llista_a_text(
                get_valor_equivalent(data, equivalencies["documentacio"])),
        }


        claus_existents = set(k for v in equivalencies.values() for k in v)
        for clau_json in data.keys():
            if clau_json not in claus_existents:
                valor = data[clau_json]
                if isinstance(valor, (list, dict)):
                    valor = json.dumps(valor, ensure_ascii=False)
                dades_planes[clau_json] = valor

        df = pd.DataFrame([dades_planes])

    except Exception as e:
        messagebox.showerror("Error JSON", f"No s'ha pogut llegir el JSON:\n{e}")
        return

    excel_path = filedialog.askopenfilename(
        title="Selecciona l'Excel on afegir les dades",
        filetypes=[("Fitxers Excel", "*.xlsx")]
    )
    if not excel_path:
        return

    try:
        book = load_workbook(excel_path)
        fulls = book.sheetnames
    except Exception as e:
        messagebox.showerror("Error Excel", f"No s'ha pogut obrir l'Excel:\n{e}")
        return

    full = simpledialog.askstring("Nom de la pestanya", f"Tria una pestanya d'aquest Excel:\n{', '.join(fulls)}")
    if full not in fulls:
        messagebox.showerror("Error", f"La pestanya '{full}' no existeix a l'Excel.")
        return

    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            startrow = book[full].max_row  # accedir directament al full
            df.to_excel(writer, sheet_name=full, index=False, header=False, startrow=startrow)

        messagebox.showinfo("Fet", "Les dades s'han afegit correctament a l'Excel.")
        text_area.delete("1.0", tk.END)
    except Exception as e:
        messagebox.showerror("Error", f"No s'han pogut afegir les dades:\n{e}")

# Interfície gràfica
root = tk.Tk()
root.title("Enganxar JSON i afegir a Excel")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

label = tk.Label(frame, text="Enganxa aquí el teu JSON:")
label.pack()

text_area = scrolledtext.ScrolledText(frame, width=80, height=20)
text_area.pack()

btn_afegir = tk.Button(frame, text="Afegir al Excel", command=afegir_json_a_excel)
btn_afegir.pack(pady=10)

root.mainloop()